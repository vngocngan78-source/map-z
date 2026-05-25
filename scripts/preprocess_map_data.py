#!/usr/bin/env python3
"""
preprocess_map_data.py
将用户的 CSV/Excel 数据转换为 map-z 所需的 JSON 格式

使用方法:
  python preprocess_map_data.py input.csv --region-col 省份 --value-col GDP
  python preprocess_map_data.py input.xlsx --region-col 城市 --value-col 销售额
"""

import json
import sys
import argparse

def normalize_region_name(name: str) -> str:
    """标准化地区名称，去掉省/市/自治区等后缀"""
    suffixes = ['省', '市', '壮族自治区', '回族自治区', '维吾尔自治区', '自治区', '特别行政区']
    for s in suffixes:
        name = name.replace(s, '')
    return name.strip()

def csv_to_mapz(input_file: str, region_col: str, value_col: str, output_file: str = None):
    """将 CSV 转换为 map-z JSON 格式"""
    try:
        import csv
        rows = []
        with open(input_file, encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if region_col not in row or value_col not in row:
                    print(f"❌ 找不到列: {region_col} 或 {value_col}")
                    print(f"   可用的列: {list(row.keys())}")
                    sys.exit(1)
                try:
                    value = float(row[value_col].replace(',', '').replace('，', ''))
                except ValueError:
                    print(f"⚠️  跳过无效数值: {row[value_col]}")
                    continue
                rows.append({
                    'name': normalize_region_name(row[region_col]),
                    'value': value
                })

        # 按 value 降序排列
        rows.sort(key=lambda x: x['value'], reverse=True)

        result = json.dumps(rows, ensure_ascii=False, indent=2)

        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(result)
            print(f"✅ 已输出到 {output_file}")
        else:
            print("\n// 复制以下内容到 REGION_DATA：")
            print(result)

        return rows

    except FileNotFoundError:
        print(f"❌ 文件不存在: {input_file}")
        sys.exit(1)

def excel_to_mapz(input_file: str, region_col: str, value_col: str, output_file: str = None):
    """将 Excel 转换为 map-z JSON 格式"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        if region_col not in headers or value_col not in headers:
            print(f"❌ 找不到列: {region_col} 或 {value_col}")
            print(f"   可用的列: {headers}")
            sys.exit(1)

        region_idx = headers.index(region_col)
        value_idx  = headers.index(value_col)

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[region_idx] is None:
                continue
            try:
                value = float(str(row[value_idx]).replace(',', ''))
            except (ValueError, TypeError):
                continue
            rows.append({
                'name': normalize_region_name(str(row[region_idx])),
                'value': value
            })

        rows.sort(key=lambda x: x['value'], reverse=True)
        result = json.dumps(rows, ensure_ascii=False, indent=2)

        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(result)
            print(f"✅ 已输出到 {output_file}")
        else:
            print("\n// 复制以下内容到 REGION_DATA：")
            print(result)

    except ImportError:
        print("❌ 需要安装 openpyxl：pip install openpyxl")
        sys.exit(1)

def generate_sample_data(scope: str = 'china') -> list:
    """生成示例数据（用于测试）"""
    china_provinces = [
        {'name':'广东','value':135000},{'name':'江苏','value':122000},
        {'name':'山东','value':92000}, {'name':'浙江','value':82000},
        {'name':'河南','value':62000}, {'name':'四川','value':57000},
        {'name':'湖北','value':55000}, {'name':'福建','value':52000},
        {'name':'湖南','value':50000}, {'name':'上海','value':47000},
        {'name':'安徽','value':46000}, {'name':'北京','value':43000},
        {'name':'河北','value':42000}, {'name':'陕西','value':33000},
        {'name':'江西','value':32000}, {'name':'重庆','value':30000},
        {'name':'辽宁','value':29000}, {'name':'云南','value':28000},
        {'name':'广西','value':26000}, {'name':'山西','value':25000},
        {'name':'天津','value':17000}, {'name':'贵州','value':21000},
        {'name':'黑龙江','value':16000},{'name':'吉林','value':13000},
        {'name':'内蒙古','value':24000},{'name':'新疆','value':18000},
        {'name':'甘肃','value':11000}, {'name':'海南','value':7000},
        {'name':'宁夏','value':5000},  {'name':'青海','value':4000},
        {'name':'西藏','value':2400},
    ]
    print(json.dumps(china_provinces, ensure_ascii=False, indent=2))
    return china_provinces

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='将数据转换为 map-z 格式')
    parser.add_argument('input', nargs='?', help='输入文件（CSV 或 Excel）')
    parser.add_argument('--region-col', default='地区', help='地区名称列名')
    parser.add_argument('--value-col',  default='数值', help='数值列名')
    parser.add_argument('--output', '-o', help='输出 JSON 文件路径')
    parser.add_argument('--sample', action='store_true', help='生成示例数据')

    args = parser.parse_args()

    if args.sample:
        generate_sample_data()
    elif args.input:
        if args.input.endswith('.xlsx') or args.input.endswith('.xls'):
            excel_to_mapz(args.input, args.region_col, args.value_col, args.output)
        else:
            csv_to_mapz(args.input, args.region_col, args.value_col, args.output)
    else:
        parser.print_help()
